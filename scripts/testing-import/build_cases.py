# -*- coding: utf-8 -*-
# 測試期間專用：把 LINE bot 匯出的原始 Excel 清洗成可匯入 Case 資料表的 JSON。
# 正式站上線後（不再靠 Excel 匯入申請資料）這支腳本連同整個 testing-import 目錄可以刪除。
#
# 用法：python3 build_cases.py <輸入.xlsx> [輸出.json，預設 cleaned_cases.json]

import openpyxl, re, sys, json
from collections import defaultdict

STANDARD = set([
    "ALTIS","ALTIS HV","C CROSS","C CROSS HV","TOWN ACE","TOWN ACE VAN","VIOS","Y CROSS",
    "ALPHARD HV","ALPHARD PHV","CAMRY","CAMRY HV","COROLLA SP","CROWN HV","GR 86","GR YARIS",
    "HILUX","LAND CRUISER","PRIUS PHV","RAV4 HV","RAV4 PHV","SIENNA HV","SUPRA",
    "URBAN CRUISER","bZ4X",
])

# 車名比對規則跑不出來的個案，手動覆寫（每次匯入新檔案時，若比對表出現新的 none/low
# confidence 項目，先跟提出檔案的人確認車型，再加進這裡）。
CAR_OVERRIDES = {
    "PU": "TOWN ACE", "PU-TSS": "TOWN ACE", "TOWN-PU": "TOWN ACE",
    "RA-H尊爵+": "RAV4 HV", "RA 尊爵+": "RAV4 HV",
    "TW-n": "TOWN ACE VAN",
}

CATEGORY_MAP = {"一般車": "一般車", "員購車": "員工車", "營業車": "營業車", "租賃車": "租賃車"}
CATEGORY_ABBR = {"一般車": "一般", "員工車": "員工", "營業車": "營業", "租賃車": "租賃"}
STATUS_MAP = {"審核通過": "APPROVED", "待審核": "PENDING_BUZHUGUAN"}

REQUIRED_HEADERS = ["申請日期", "所別", "課別", "領牌名稱", "訂單編號", "特案類別", "車名",
                    "所課支援金", "金牌", "折讓總額", "特案支援金額", "備註", "審核狀態"]


def norm_car(raw):
    if raw in CAR_OVERRIDES:
        return CAR_OVERRIDES[raw]
    s = raw.upper()
    s = s.replace("Ｈ", "H").replace("－", "-").replace("–", "-").replace("—", "-")
    s = re.sub(r'[\s／/]+', '', s)
    is_phv = "PHEV" in s or "PHV" in s
    is_hv = (not is_phv) and bool(
        re.search(r'HEV|HYBRID|油電|HV|-H$|H$', s)
        or re.match(r'^(CC|CROSS|CCROSS|AL|ALTIS|CM|CAMRY|R4|RA|RAV4)-?H(?![A-Z])', s)
    )
    if "GRYARIS" in s: return "GR YARIS"
    if "GR86" in s: return "GR 86"
    if "YARISCROSS" in s or "YCROSS" in s or s.startswith("YC"): return "Y CROSS"
    if "YARIS" in s: return "Y CROSS"
    if "ALPHARD" in s: return "ALPHARD PHV" if is_phv else "ALPHARD HV"
    if "CROWN" in s: return "CROWN HV"
    if "SIENNA" in s: return "SIENNA HV"
    if "PRIUS" in s: return "PRIUS PHV"
    if "URBANCRUISER" in s: return "URBAN CRUISER"
    if "LANDCRUISER" in s or "LC300" in s or "LC250" in s: return "LAND CRUISER"
    if "HILUX" in s: return "HILUX"
    if "SUPRA" in s: return "SUPRA"
    if "BZ4X" in s: return "bZ4X"
    if "COROLLASP" in s or "CLSP" in s: return "COROLLA SP"
    if "RAV4" in s or "RAV-4" in s or re.match(r'^R4', s):
        return "RAV4 PHV" if is_phv else "RAV4 HV"
    if "TOWNACE" in s or "TOWN-ACE" in s or "TOWN.ACE" in s or s.startswith("TA") or s.startswith("TAV") or s.startswith("TVN") or "ACE" in s:
        return "TOWN ACE VAN" if ("VAN" in s or "VEN" in s) else "TOWN ACE"
    if "VAN" in s or "VEN" in s or s.startswith("TWN"): return "TOWN ACE VAN"
    if s == "PU" or s.startswith("PU-") or s.endswith("-PU") or s == "PUTSS": return "TOWN ACE"
    if s.startswith("CAMRY") or s.startswith("CM") or "CMARY" in s: return "CAMRY HV" if is_hv else "CAMRY"
    if s.startswith("VIOS"): return "VIOS"
    if s.startswith("ALTIS") or s.startswith("AL"): return "ALTIS HV" if is_hv else "ALTIS"
    if s.startswith("COROLLACROSS"): return "C CROSS HV" if is_hv else "C CROSS"
    if s.startswith("CCROSS") or s.startswith("CC") or s.startswith("CROSS"): return "C CROSS HV" if is_hv else "C CROSS"
    if re.match(r'^RA-?H|^RA\d|^RA尊|^RA旗', s): return "RAV4 HV"
    return None


def norm_dept(raw):
    s = str(raw).strip()
    if not s:
        return None
    if s[0] in ('1', '2', '3'):
        return s[0]
    if s[0] in ('一', 'ㄧ'):
        return '1'
    if s[0] == '二':
        return '2'
    if s[0] == '三':
        return '3'
    return None


def main():
    if len(sys.argv) < 2:
        print("用法: python3 build_cases.py <輸入.xlsx> [輸出.json]")
        sys.exit(1)
    in_path = sys.argv[1]
    out_path = sys.argv[2] if len(sys.argv) > 2 else "cleaned_cases.json"

    wb = openpyxl.load_workbook(in_path, data_only=True)
    ws = wb.worksheets[0]
    header = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {}
    for h in REQUIRED_HEADERS:
        if h not in header:
            print(f"錯誤：找不到欄位「{h}」，原始欄位有：{header}")
            sys.exit(1)
        idx[h] = header.index(h)

    by_order = {}
    skipped_no_order = 0
    skipped_bad_field = []
    unmatched_cars = defaultdict(int)

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[idx["申請日期"]] is None:
            continue  # 尾端空白列
        order = str(row[idx["訂單編號"]]).strip() if row[idx["訂單編號"]] else ''
        if not re.match(r'^D\d{12}$', order):
            skipped_no_order += 1
            continue
        date = row[idx["申請日期"]]
        if order in by_order and by_order[order][0] >= date:
            continue  # 同訂單編號重複出現（案件被編輯過）：留最新一筆
        by_order[order] = (date, row)

    records = []
    for order, (date, row) in by_order.items():
        store = str(row[idx["所別"]]).strip()
        dept = norm_dept(row[idx["課別"]])
        cat_raw = str(row[idx["特案類別"]]).strip() if row[idx["特案類別"]] else None
        cat = CATEGORY_MAP.get(cat_raw)
        car_raw = str(row[idx["車名"]]).strip() if row[idx["車名"]] else None
        car = norm_car(car_raw) if car_raw else None
        status = STATUS_MAP.get(str(row[idx["審核狀態"]]).strip()) if row[idx["審核狀態"]] else None
        if not (re.match(r'^D\d{2}$', store) and dept and cat and car and status):
            if car_raw and not car:
                unmatched_cars[car_raw] += 1
            skipped_bad_field.append((order, store, dept, cat_raw, car_raw, row[idx["審核狀態"]]))
            continue
        records.append({
            "orderNo": order,
            "storeCode": store,
            "deptCode": dept,
            "plateName": str(row[idx["領牌名稱"]]).strip() if row[idx["領牌名稱"]] else "",
            "category": cat,
            "carModel": car,
            "description": str(row[idx["備註"]]).strip() if row[idx["備註"]] else "",
            "subsidyDeptCourse": int(row[idx["所課支援金"]]) if row[idx["所課支援金"]] else 0,
            "goldMedal": int(row[idx["金牌"]]) if row[idx["金牌"]] else 0,
            "silverMedal": 0,  # 原始檔案沒有銀牌欄位
            "discountTotal": int(row[idx["折讓總額"]]) if row[idx["折讓總額"]] else 0,
            "specialSubsidy": int(row[idx["特案支援金額"]]) if row[idx["特案支援金額"]] else 0,
            "status": status,
            "submittedAt": date.isoformat(),
            "month": f"{date.year:04d}-{date.month:02d}",
        })

    records.sort(key=lambda x: x["submittedAt"])
    seq = defaultdict(int)
    for rec in records:
        key = (rec["storeCode"], rec["category"])
        seq[key] += 1
        rec["categoryNo"] = f'{CATEGORY_ABBR[rec["category"]]}{seq[key]}'

    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2)

    print(f"跳過（訂單編號缺失/格式不對）: {skipped_no_order}")
    print(f"跳過（所別/課別/類別/車名/狀態比對失敗）: {len(skipped_bad_field)}")
    for s in skipped_bad_field:
        print(" ", s)
    if unmatched_cars:
        print("無法比對的車名（需要加進 CAR_OVERRIDES 或補比對規則）：")
        for k, v in unmatched_cars.items():
            print(f"  {v:3d}  {k!r}")
    print(f"乾淨資料筆數: {len(records)} -> {out_path}")


if __name__ == "__main__":
    main()
